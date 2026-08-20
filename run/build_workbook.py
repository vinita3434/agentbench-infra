#!/usr/bin/env python3
"""Build results/agentbench.xlsx from the master table.

Tabs:
  Master              every episode, one row, with a stable episode id
  Common_All          tasks every model in the set has attempted
  <Frontier>_vs_<Open>  one tab per frontier x open-weight pair sharing tasks

Comparison tabs are generated, not enumerated, so adding a model adds its tabs
automatically. Only **frontier vs open-weight** pairs are produced: the question
this benchmark exists to answer is whether a self-hostable model can stand in
for a paid API, so opus-vs-sonnet (two frontier) and kimi-vs-deepseek (two open)
are noise. Add a model to MODEL_CLASS and its tabs appear on the next run.

    python run/build_workbook.py
"""
from __future__ import annotations

import csv
import pathlib
import sys
from collections import defaultdict

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
CSV_PATH = REPO_ROOT / "results" / "master_table.csv"
XLSX_PATH = REPO_ROOT / "results" / "agentbench.xlsx"

# Task-id sets, used to label which benchmark each episode came from.
SWEBENCH_IDS = {p.stem for p in (REPO_ROOT / "tasks" / "swebench_lite" / "data").glob("*.json")}
POLYBENCH_IDS = {p.stem for p in (REPO_ROOT / "tasks" / "data").glob("*.json")}

# Which side of the comparison a model sits on. Keyed by the provider prefix of
# the model string, so a new model from a known provider is classified without
# an edit; unknown providers default to "open" and are flagged on stdout.
PROVIDER_CLASS = {
    "anthropic": "frontier",
    "openai": "frontier",
    "google": "frontier",
    "x-ai": "frontier",
    "moonshotai": "open",
    "qwen": "open",
    "deepseek": "open",
    "mistralai": "open",
    "meta-llama": "open",
    "zai-org": "open",
    "selfhosted": "open",     # anything served locally via SGLang
}


def model_class(model: str) -> str:
    provider = model.split("/")[0] if "/" in model else ""
    return PROVIDER_CLASS.get(provider, "open")


def short(model: str) -> str:
    return model.split("/")[-1]


def abbrev(name: str) -> str:
    """Shorten for a 31-char sheet title without mangling the model name.

    Only whole-token prefixes are dropped -- a naive substring replace turned
    "qwen3-coder-plus" into "qwen3r-plus".
    """
    for prefix in ("claude-", "moonshotai-"):
        if name.startswith(prefix):
            name = name[len(prefix):]
    return name[:14]


def active_share(part, other) -> str:
    """"5m12s (62%)" where the percentage is of GPU+CPU only.

    Overhead is deliberately left out of the denominator: it is Pi's own
    processing and stream I/O, not work either side did, and including it makes
    both shares shrink for reasons unrelated to the model. With it excluded the
    two percentages sum to 100 and answer the question directly -- of the time
    actually spent working, how much was the model thinking versus tools running.
    """
    part = num(part) or 0.0
    total = part + (num(other) or 0.0)
    if not total:
        return hms(part)
    return f"{hms(part)} ({100 * part / total:.0f}%)"


def hms(seconds) -> str:
    if not seconds:
        return ""
    m, s = divmod(int(float(seconds)), 60)
    return f"{m}m{s:02d}s" if m else f"{s}s"


def num(value, cast=float):
    try:
        return cast(value)
    except (TypeError, ValueError):
        return None


def load():
    if not CSV_PATH.exists():
        sys.exit(f"no {CSV_PATH} -- run python run/master_table.py first")
    rows = list(csv.DictReader(CSV_PATH.open()))
    # Stable episode id: model slug + task + timestamp. Deterministic across
    # rebuilds, so a row keeps its id when new runs are appended.
    for r in rows:
        # task + model + attempt. Stable across rebuilds (unlike a timestamp),
        # and unique when the same model retries the same task -- which is the
        # whole reason attempts are tracked separately per model.
        r["episode_id"] = (
            f"{r['task_id']}|{short(r['model'])}|a{r.get('model_attempt_n', 1)}"
        )
        r["class"] = model_class(r["model"])
        r["dataset"] = ("SWE-bench Lite" if r["task_id"] in SWEBENCH_IDS
                        else "SWE-PolyBench" if r["task_id"] in POLYBENCH_IDS else "?")
        # The harness is Pi, unmodified, unless a context-strategy extension was
        # threaded in. Recording it per episode is what makes round two -- hold
        # the model fixed, vary the harness -- readable off the same sheet.
        ext = (r.get("pi_extension") or "").strip()
        r["harness"] = f"pi + {pathlib.Path(ext).stem}" if ext else "pi (baseline)"
    rows.sort(key=lambda r: (r["model"], r["task_id"], r.get("timestamp_utc") or ""))
    return rows


HEADER_FILL = "FF1F3864"
BAND_FILL = "FFF2F2F2"


def style_sheet(ws, header_rows=1, widths=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    for row in range(1, header_rows + 1):
        for cell in ws[row]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
    ws.freeze_panes = ws.cell(row=header_rows + 1, column=1)
    if ws.max_row > header_rows:
        ws.auto_filter.ref = (
            f"A{header_rows}:{get_column_letter(ws.max_column)}{ws.max_row}"
        )
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        longest = max(
            (len(str(ws.cell(row=r, column=i).value or ""))
             for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[letter].width = min(
            max((widths or {}).get(i, 0), longest + 2, 9), 52)


def sheet_master(wb, rows):
    ws = wb.create_sheet("Master")
    ws.append(["Episode ID", "Task", "Dataset", "Harness", "Model", "Attempt",
               "Wall Clock Time", "Cost (USD)", "GPU (reasoning) time",
               "CPU (tool exec) time", "Resolved", "Difficulty", "Turns"])
    for r in rows:
        ws.append([
            r["episode_id"], r["task_id"], r["dataset"], r["harness"],
            short(r["model"]),
            f"{r.get('model_attempt_n', 1)}/{r.get('model_attempts_total', 1)}",
            # Agent span, not run_record.wall_seconds: a bash tool that leaves a
            # background process holding stdout keeps the shell waiting long
            # after the agent finished, and that is not work.
            hms(r.get("agent_span_s")),
            num(r["cost_usd"]),
            active_share(r.get("model_time_s"), r.get("tool_time_s")),
            active_share(r.get("tool_time_s"), r.get("model_time_s")),
            r["resolved"], r["difficulty"], num(r["turns"], int),
        ])
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '"$"#,##0.0000'
    style_sheet(ws)
    return ws


def _comparison_rows(rows, models):
    """Tasks every model in `models` has attempted, latest attempt each."""
    per = defaultdict(dict)
    for r in rows:
        m = short(r["model"])
        if m in models:
            per[r["task_id"]][m] = r      # rows are timestamp-sorted; last wins
    return {t: v for t, v in per.items() if len(v) == len(models)}


def sheet_comparison(wb, title, rows, models):
    """One row per shared task, every model's numbers side by side.

    CPVT is cost-per-verified-task: the episode's cost when it resolved, blank
    when it did not. A failed episode still costs money but yields no verified
    task, so averaging its spend in would understate what each success costs.
    The TOTAL row divides all spend by the number verified -- the honest
    aggregate.

    GPU and tool percentages are shares of GPU+CPU, overhead excluded, so they
    sum to 100 and describe how working time was split.
    """
    shared = _comparison_rows(rows, models)
    if not shared:
        return None
    ws = wb.create_sheet(title[:31])

    names = [short(m) for m in models]
    header = ["Episode ID"]
    for label in ("Wall Clock", "CPVT", "Turns", "GPU time", "Tool exec"):
        header += [f"{n} {label}" for n in names]
    header += ["Dataset", "Difficulty"] + [f"{n} Resolved" for n in names]
    ws.append(header)

    order = {"easy": 0, "medium": 1, "hard": 2, "unrated": 3}
    for task in sorted(shared, key=lambda t: (
            order.get(list(shared[t].values())[0]["difficulty"], 9), t)):
        v = shared[task]
        rs = [v[m] for m in models]
        line = [task]
        line += [hms(r.get("agent_span_s")) for r in rs]
        line += [num(r["cost_usd"]) if r["resolved"] == "yes" else None for r in rs]
        line += [num(r["turns"], int) for r in rs]
        line += [active_share(r.get("model_time_s"), r.get("tool_time_s")) for r in rs]
        line += [active_share(r.get("tool_time_s"), r.get("model_time_s")) for r in rs]
        first = rs[0]
        line += [first["dataset"], first["difficulty"]] + [r["resolved"] for r in rs]
        ws.append(line)

    def agg(model_key):
        sub = [v[model_key] for v in shared.values()]
        won = [r for r in sub if r["resolved"] == "yes"]
        spend = sum(num(r["cost_usd"]) or 0 for r in sub)
        scored = [r for r in sub if r["resolved"] in ("yes", "no")]
        return dict(
            wall=sum(num(r.get("agent_span_s")) or 0 for r in sub),
            cpvt=(spend / len(won)) if won else None,
            turns=sum(num(r["turns"], int) or 0 for r in sub),
            gpu=sum(num(r.get("model_time_s")) or 0 for r in sub),
            cpu=sum(num(r.get("tool_time_s")) or 0 for r in sub),
            res=f"{len(won)}/{len(scored)}")
    totals = [agg(m) for m in models]
    row = ["TOTAL"]
    row += [hms(t["wall"]) for t in totals]
    row += [t["cpvt"] for t in totals]
    row += [t["turns"] for t in totals]
    row += [active_share(t["gpu"], t["cpu"]) for t in totals]
    row += [active_share(t["cpu"], t["gpu"]) for t in totals]
    row += [f"{len(shared)} tasks", ""] + [t["res"] for t in totals]
    ws.append(row)

    from openpyxl.styles import Font
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    cpvt_start = 2 + len(models)
    for col in range(cpvt_start, cpvt_start + len(models)):
        for r in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in r:
                cell.number_format = '"$"#,##0.0000'
    style_sheet(ws, header_rows=1)
    return ws


def main():
    try:
        from openpyxl import Workbook
    except ImportError:
        sys.exit("pip install openpyxl")

    rows = load()
    models = sorted({short(r["model"]) for r in rows})
    cls = {short(r["model"]): r["class"] for r in rows}
    frontier = [m for m in models if cls[m] == "frontier"]
    openw = [m for m in models if cls[m] == "open"]

    wb = Workbook()
    wb.remove(wb.active)
    sheet_master(wb, rows)
    print(f"Master: {len(rows)} episodes, {len(models)} models")

    # Tasks a whole group of models has attempted -- the fully controlled
    # comparison. Try all models first; if nothing is shared by all of them,
    # fall back to the largest subset that does share tasks, so the tab is
    # useful rather than empty.
    from itertools import combinations
    best = None
    for size in range(len(models), 2, -1):
        for combo in combinations(models, size):
            shared = _comparison_rows(rows, list(combo))
            if shared and (best is None or len(shared) > len(best[1])):
                best = (list(combo), shared)
        if best:
            break
    if best:
        ws = sheet_comparison(wb, "Common", rows, best[0])
        print(f"Common: {len(best[1])} tasks shared by {', '.join(best[0])}")
    else:
        print("Common: no task attempted by 3+ models yet")

    # Frontier x open only. A new open model pairs against every frontier model
    # automatically; frontier-vs-frontier and open-vs-open are never emitted.
    for f in frontier:
        for o in openw:
            title = f"{abbrev(f)}_vs_{abbrev(o)}"
            ws = sheet_comparison(wb, title, rows, [f, o])
            if ws:
                print(f"{ws.title}: {ws.max_row - 3} shared tasks")

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"\n-> {XLSX_PATH}")
    print(f"   tabs: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
